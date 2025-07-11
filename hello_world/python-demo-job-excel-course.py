#通过xlrd和xlsxwriter实现excel表格文件中的多sheet内容的合并
import xlrd,xlsxwriter,openpyxl
from openpyxl.utils import get_column_letter

'''
实现方案
1.抽提单个文件方法
2.方法返回指定excel文件中的所有sheet及sheet对应数据的二元元组
2.1 剔除不同sheet的表头名
3.拿到文件中的数据后，往指定文件中写入。自定义sheet名
4.实现最终的合并sheet写入效果
'''
#根据文件路径获取到指定的excel文件
def read_excel(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        return read_excel_sheet_data(workbook)
    except Exception as e:
        print(f"{file_path}下的文件打开出错{e}")

#读取文件的所有sheet，以及sheet对应的table内容，然后返回
def read_excel_sheet_data(workbook):
    # 获取到所有活动的sheet
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
        # 根据名称获取到sheet
        sheet = workbook[sheet_name]
        # 获取到当前sheet下的所有行，遍历后存放到data中,[min_row=2]从第二行开始读取，即跳过表头
        data = [row for row in sheet.iter_rows(values_only=True)]
        # 将指定sheet_name下的数据所有行数据，存放到sheets_data中
        sheets_data[sheet_name] = data
        '''
        最终存放的数据长这样（类似）：
        {
        "Sheet1": [["姓名", "年龄"], ["张三", 25]],
        "Sheet2": [["产品", "价格"], ["手机", 5000]]
        }
        '''
    workbook.close()
    return sheets_data

def combine_sheet_data(sheet_data,sheet_name,file_path):
    workbook = openpyxl.Workbook()
    #删除创建的默认sheet
    workbook.remove(workbook.active)
    #创建一个新sheet
    sheet = workbook.create_sheet(sheet_name)
    #遍历sheet_data,填充数据
    for sheet_name,data in sheet_data.items():
        for row in data:
            sheet.append(list(row) if isinstance(row, (tuple, set)) else row)
    #自动调整列宽
    # 自动调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width
    # 保存工作簿
    workbook.save(file_path)
    workbook.close()
    print(f"成功将数据写入文件 '{file_path}'")

if __name__ == "__main__":
    # 读取excel文件
    input_file_1 = "/demo/testFile/test-excel.xlsx"
    print(f"读取文件 '{input_file_1}'...")
    multi_excel_data_1 = read_excel(input_file_1)
    input_file_2 = "/demo/testFile/test-excel-2.xlsx"
    print(f"读取文件 '{input_file_2}'...")
    multi_excel_data_2 = read_excel(input_file_2)
    multi_excel_data = {**multi_excel_data_1,**multi_excel_data_2}
    print(f"\n将数据写入文件 'test-excel-multi-course'....")
    output_file = "/demo/testFile/test-excel-multi-output-course.xlsx"
    combine_sheet_data(multi_excel_data,"合并数据-course",output_file)