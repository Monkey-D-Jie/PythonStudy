#通过第三方库，读写excel表格，来自豆包
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment

def read_excel_single_sheet(file_path):
    #读取excel并返回数据
    try:
        workbook = openpyxl.load_workbook(file_path)
        data = read_excel_single_sheet_data(workbook)  # 传递 workbook 对象
        workbook.close()
        return data
    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
        return []
    except Exception as e:
        print(f"错误：读取文件时发生错误-{e}")
        return None
def read_excel_multi_sheet(file_path):
    #读取excel并返回数据
    try:
        workbook = openpyxl.load_workbook(file_path)
        data = read_excel_multi_sheet_data(workbook)
        workbook.close()
        return data
    except FileNotFoundError:
        print(f"错误：找不到文件 '{file_path}'")
        return []
    except Exception as e:
        print(f"错误：读取文件时发生错误-{e}")
        return None
def read_excel_single_sheet_data(workbook):
    data = []
    # 获取到活动的sheet-获取到单一sheet
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def read_excel_multi_sheet_data(workbook):
    # 获取到所有活动的sheet
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
        # 根据名称获取到sheet
        sheet = workbook[sheet_name]
        # 获取到当前sheet下的所有行，遍历后存放到data中
        data = [row for row in sheet.iter_rows(values_only=True)]
        # 将指定sheet_name下的数据所有行数据，存放到sheets_data中
        sheets_data[sheet_name] = data
        '''
        最终存放的数据长这样（类似）：
        {
        "Sheet1": [["姓名", "年龄"], ["张三", 25]],
        "Sheet2": [["产品", "价格"], ["手机", 5000]]
        }
        '''
    workbook.close()
    return sheets_data

def write_excel(file_path,sheet_data,sheet_name="Sheet1",headers=None,single_sheet=False):
    #将谁写入excel文件
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title=sheet_name
        #如果提供了表头，则写入表头
        if headers:
            sheet.append(headers)
            for cell in sheet[1]:
                 cell.font=Font(bold=True)
                 cell.alignment=Alignment(horizontal='center')
        #写入数据
        if single_sheet:
            for row in sheet_data:
                sheet.append(row)
        else:
            workbook.remove(workbook.active)  # 删除默认创建的 sheet
            #指定统一的sheet名
            combine_sheet_name = "合并sheet"
            # 创建新的sheet
            sheet = workbook.create_sheet(combine_sheet_name)
            for sheet_name,data in sheet_data.items():
                if headers:  # 如果提供了通用 headers
                    sheet.append(headers)
                    # ...（保持原有样式代码）
                # 打印每个 sheet 的前两行
                for row in data:  # 确保 data 是二维结构
                    sheet.append(list(row) if isinstance(row, (tuple, set)) else row)

        #自动调整列宽
        for column in sheet.columns:
            max_length =0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length=len(str(cell.value))
                except:
                    pass
            adjusted_width=(max_length+2)
            sheet.column_dimensions[column_letter].width=adjusted_width
        #保存工作簿
        workbook.save(file_path)
        workbook.close()
        print(f"成功将数据写入文件 '{file_path}'")
    except Exception as e:
        print(f"错误：写入文件时发生错误-{e}")

if __name__ == "__main__":
    #读取excel文件
    input_file= "/demo/testFile/test-excel.xlsx"
    print(f"读取文件 '{input_file}'...")
    single_excel_data=read_excel_single_sheet(input_file)
    if(single_excel_data):
        #打印前几行数据
        print("数据前几行内容：")
        for i,row in enumerate(single_excel_data[:5]):
            print(f"行 {i+1}: {row}")
        #写入excel文件
        #单个sheet写入
        output_file= "/testFile/test-excel-single-output.xlsx"
        # headers=["a","b","c"]
        print(f"\n将数据写入文件 'test-excel-single-output'....")
        write_excel(output_file,single_excel_data,"single-output",None,True)
    multi_excel_data = read_excel_multi_sheet(input_file)

    if (multi_excel_data):
        print(f"\n将数据写入文件 'test-excel-multi-output'....")
        output_file = "/testFile/test-excel-multi-output.xlsx"
        # headers=["a","b","c"]
        write_excel(output_file, multi_excel_data, "multi-output", None, False)
